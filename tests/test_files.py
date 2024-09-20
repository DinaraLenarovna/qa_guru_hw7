import csv
import zipfile
from openpyxl import load_workbook
from pypdf import PdfReader
from script_os import ZIP_DIR


def test_xlsx():
    with zipfile.ZipFile(ZIP_DIR, 'r') as zip_file:
        with zip_file.open("school_example.xlsx") as excel_file:
            wb = load_workbook(filename=excel_file)
            sheet = wb.active
            value = sheet.cell(row=4, column=1).value
            name = 'Соня'
            assert name in value


def test_csv():
    with zipfile.ZipFile(ZIP_DIR, 'r') as zip_file:
        with zip_file.open("csv_example.csv") as csv_file:
            content = csv_file.read().decode(
                'utf-8-sig')  # читаем содержимое файла и декодируем его если в файле есть символы не из английского алфавита
            csvreader = list(csv.reader(content.splitlines()))  # читаем содержимое файла и преобразуем его в список
            second_row = csvreader[1]  # получаем вторую строку
            assert second_row[0] == 'batman'
            assert second_row[1] == 'uses technology'


def test_pdf():
    with zipfile.ZipFile(ZIP_DIR, 'r') as zip_file:
        with zip_file.open("pdf_test.pdf") as pdf_file:
            reader = PdfReader(pdf_file)
            page = reader.pages[0]
            text = page.extract_text()
            assert 'Башкирский химический журнал' in text
